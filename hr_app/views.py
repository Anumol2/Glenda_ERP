from django.shortcuts import render,redirect,get_object_or_404,reverse
from register_app.models import CustomUser, MenuPermissions
from hr_app.models import EmployeeDetails,RequestLeave,Payroll
from Glenda_App.models import Menu
from .forms import EmployeeDetailsForm,LeaveRequestForm,PayrollForm
from django.db.models import Q
from django.http import HttpResponse
from django.contrib import messages
from django.template.loader import render_to_string, get_template
from xhtml2pdf import pisa
from openpyxl import Workbook
from openpyxl.styles import Font,Alignment
from django.contrib.auth.decorators import login_required
from django.core.paginator import Paginator
from django.views.decorators.csrf import csrf_exempt
from django.http import JsonResponse
import csv
from decimal import Decimal  # For setting a default fallback value of zero
from django.utils import timezone
def Employee_list(request):
    use = request.user  # Get the current user

    # Get user's permissions
    user_permissions = MenuPermissions.objects.filter(user=use).select_related('menu', 'sub_menu')

    # Create a dictionary to hold menus and their allocated submenus
    allocated_menus = {}
    for perm in user_permissions:
        if perm.menu not in allocated_menus:
            allocated_menus[perm.menu] = []
        allocated_menus[perm.menu].append(perm.sub_menu)


    # Fetch all users with is_staff=True
    users = CustomUser.objects.filter(is_staff=True,is_superuser=False)

    for user in users:

        user.details_added = EmployeeDetails.objects.filter(user=user).exists()


    # Pass the filtered HR employee details and menus to the context
    context = {
        'allocated_menus': allocated_menus,
        'users':users
    }

    return render(request, 'hr/Employee_list.html', context)

def AddDetails(request, id):

    use = request.user  # Get the current user

    # Get user's permissions
    user_permissions = MenuPermissions.objects.filter(user=use).select_related('menu', 'sub_menu')

    # Create a dictionary to hold menus and their allocated submenus
    allocated_menus = {}
    for perm in user_permissions:
        if perm.menu not in allocated_menus:
            allocated_menus[perm.menu] = []
        allocated_menus[perm.menu].append(perm.sub_menu)

    if request.method == 'POST':
        form = EmployeeDetailsForm(request.POST, request.FILES)

        emergency_contact_number = request.POST.get('emergency_contact_number')
        aadhar_no = request.POST.get('aadhar_no')
        employee_esi_no = request.POST.get('employee_esi_no')
        pf_no = request.POST.get('pf_no')

        # Validation checks
        if len(str(emergency_contact_number)) != 10:
            messages.warning(request, 'Contact number must be exactly 10 digits.')

        if len(str(aadhar_no)) != 12:
            messages.warning(request, 'Aadhar number must be exactly 12 digits.')

        if len(str(employee_esi_no)) != 20:
            messages.warning(request, 'ESI number must be exactly 20 digits.')

        if len(str(pf_no)) != 17:
            messages.warning(request, 'PF number must be exactly 17 digits.')

        if EmployeeDetails.objects.filter(aadhar_no=aadhar_no).exists():
            messages.warning(request, 'Aadhar number already exists.')
            return render(request, 'hr/add_employee_details.html', {'form': form, 'allocated_menus': allocated_menus})

        if EmployeeDetails.objects.filter(emergency_contact_number=emergency_contact_number).exists():
            messages.warning(request, 'Phone number already exists.')
            return render(request, 'hr/add_employee_details.html', {'form': form, 'allocated_menus': allocated_menus})

        if EmployeeDetails.objects.filter(employee_esi_no=employee_esi_no).exists():
            messages.warning(request, 'ESI Number already exists.')
            return render(request, 'hr/add_employee_details.html', {'form': form, 'allocated_menus': allocated_menus})

        if EmployeeDetails.objects.filter(pf_no=pf_no).exists():
            messages.warning(request, 'PF number already exists.')
            return render(request, 'hr/add_employee_details.html', {'form': form, 'allocated_menus': allocated_menus})

        if form.is_valid():
            employee = form.save(commit=False)  # Do not save yet
            employee.user_id = id  # Associate the user
            employee.save()  # Now save it

            return redirect('Employee_list')  # Redirect to employee list after successful submission
    else:
        form = EmployeeDetailsForm()

    return render(request, 'hr/add_employee_details.html', {'form': form, 'allocated_menus': allocated_menus})

def view_employee_profile(request,id):
    use = request.user  # Get the current user

    # Get user's permissions
    user_permissions = MenuPermissions.objects.filter(user=use).select_related('menu', 'sub_menu')

    # Create a dictionary to hold menus and their allocated submenus
    allocated_menus = {}
    for perm in user_permissions:
        if perm.menu not in allocated_menus:
            allocated_menus[perm.menu] = []
        allocated_menus[perm.menu].append(perm.sub_menu)

    view = get_object_or_404(EmployeeDetails,user_id=id)

    return render(request,'hr/view_employee_profile.html',{'view':view,'allocated_menus':allocated_menus})

def update_employee_details(request,id):
    use = request.user  # Get the current user

    # Get user's permissions
    user_permissions = MenuPermissions.objects.filter(user=use).select_related('menu', 'sub_menu')

    # Create a dictionary to hold menus and their allocated submenus
    allocated_menus = {}
    for perm in user_permissions:
        if perm.menu not in allocated_menus:
            allocated_menus[perm.menu] = []
        allocated_menus[perm.menu].append(perm.sub_menu)

    view = get_object_or_404(EmployeeDetails, user_id=id)

    if request.method == 'POST':
        form = EmployeeDetailsForm(request.POST,instance=view)

        if form.is_valid():
            form.save()
            return redirect('Employee_list')

    else:
        form = EmployeeDetailsForm(request.POST, instance=view)
    return render(request, 'hr/update_employee_details.html', {'view':view, 'allocated_menus': allocated_menus,'form':form})

def delete_employee_details(request, id):
    use = request.user  # Get the current user

    # Get user's permissions
    user_permissions = MenuPermissions.objects.filter(user=use).select_related('menu', 'sub_menu')

    # Create a dictionary to hold menus and their allocated submenus
    allocated_menus = {}
    for perm in user_permissions:
        if perm.menu not in allocated_menus:
            allocated_menus[perm.menu] = []
        allocated_menus[perm.menu].append(perm.sub_menu)

    # Fetch the object or return a 404 error if not found
    dtl = get_object_or_404(EmployeeDetails, id=id)

    if request.method == "POST":
        dtl.delete()
        return redirect('Employee_list')

    # Render the confirmation page for GET requests
    return render(request, 'hr/delete_employee_details.html', {'dtl': dtl,'allocated_menus':allocated_menus})


def employee_search_and_export(request):
    use = request.user  # Get the current user

    # Get user's permissions
    user_permissions = MenuPermissions.objects.filter(user=use).select_related('menu', 'sub_menu')

    # Create a dictionary to hold menus and their allocated submenus
    allocated_menus = {}
    for perm in user_permissions:
        if perm.menu not in allocated_menus:
            allocated_menus[perm.menu] = []
        allocated_menus[perm.menu].append(perm.sub_menu)

    # Start with an empty query to show all staff members by default
    employee_list = CustomUser.objects.filter(is_staff=True, is_superuser=False)

    search_query = request.GET.get('search_query', '').strip()
    export_format = request.GET.get('export', '')

    if search_query:
        filters = Q(is_staff=True, is_superuser=False)
        if search_query.isdigit():
            filters &= Q(phone_number__icontains=search_query)
        else:
            filters &= Q(name__icontains=search_query)

        employee_list = CustomUser.objects.filter(filters)

    # Check for export format (either 'csv' or 'pdf')
    if export_format == 'excel':
        wb = Workbook()
        ws = wb.active
        ws.title = 'Employee List'
        ws.merge_cells('A1:D1')
        ws['A1'] = 'Employee List'
        headline_font = Font(bold=True, size=14)
        ws['A1'].font = headline_font
        ws['A1'].alignment = Alignment(horizontal='center')

        headers = ['Name', 'Email', 'Phone Number']
        ws.append(headers)

        header_font = Font(bold=True)
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=2, column=col_num)
            cell.font = header_font

        for employee in employee_list:
            ws.append([employee.name, employee.email, employee.phone_number])

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="employee_list.xlsx"'

        wb.save(response)
        return response

    elif export_format == 'pdf':
        # PDF export
        template = get_template('hr/employee_details_pdf.html')  # A separate template for PDF
        context = {'users': employee_list, 'allocated_menus': allocated_menus}
        html = template.render(context)

        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="employee_list.pdf"'

        pdf_status = pisa.CreatePDF(html, dest=response)

        if pdf_status.err:
            return HttpResponse('Error generating PDF', status=500)

        return response

    # Normal page rendering (with search functionality)

    for user in employee_list:

        user.details_added = EmployeeDetails.objects.filter(user=user).exists()

    context = {
        'users': employee_list,  # Filtered employees list
        'allocated_menus': allocated_menus,
    }

    return render(request, 'hr/Employee_list.html', context)

def view_leave_list(request):
    # Fetch all employees
    employees = EmployeeDetails.objects.all()

    # Prepare a list to hold employee leave requests
    leave_requests = []

    for employee in employees:
        # Get all leave requests for each employee
        requests = RequestLeave.objects.filter(employee=employee)
        leave_requests.append({
            'employee': employee,
            'requests': requests  # This will be a QuerySet of leave requests
        })

    # Apply pagination to leave_requests list
    paginator = Paginator(leave_requests, 10)  # Show 10 employees per page
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    use = request.user  # Get the current user

    # Get user's permissions
    user_permissions = MenuPermissions.objects.filter(user=use).select_related('menu', 'sub_menu')

    # Create a dictionary to hold menus and their allocated submenus
    allocated_menus = {}
    for perm in user_permissions:
        if perm.menu not in allocated_menus:
            allocated_menus[perm.menu] = []
        allocated_menus[perm.menu].append(perm.sub_menu)

    # Pass the paginated leave_requests and menus to the context
    context = {
        'allocated_menus': allocated_menus,
        'page_obj': page_obj  # Include the page object in the context
    }

    return render(request, 'hr/leave_list.html', context)


@csrf_exempt
def approve_leave_request(request):
    if request.method == 'POST':
        user_id = request.POST.get('employee_user_id')
        try:
            leave_request = RequestLeave.objects.get(id=user_id)
            leave_request.approval_status = 'APPROVED'  # Update status to 'approved'
            leave_request.save()
            return JsonResponse({'success': True})
        except RequestLeave.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'Leave request not found'})

@csrf_exempt
def reject_leave_request(request):
    if request.method == 'POST':
        user_id = request.POST.get('employee_user_id')
        rejection_reason = request.POST.get('rejection_reason')
        try:
            leave_request = RequestLeave.objects.get(id=user_id)
            leave_request.approval_status = 'REJECTED'  # Update status to 'rejected'
            leave_request.rejection_reason = rejection_reason
            leave_request.save()
            return JsonResponse({'success': True})
        except RequestLeave.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'Leave request not found'})



def my_profile(request):
    user = request.user
    leave_requests = RequestLeave.objects.filter(employee__user=user).order_by('-start_date')  # Order by date
    use = request.user  # Get the current user

    # Get user's permissions
    user_permissions = MenuPermissions.objects.filter(user=use).select_related('menu', 'sub_menu')

    # Create a dictionary to hold menus and their allocated submenus
    allocated_menus = {}
    for perm in user_permissions:
        if perm.menu not in allocated_menus:
            allocated_menus[perm.menu] = []
        allocated_menus[perm.menu].append(perm.sub_menu)

    # Pagination logic
    paginator = Paginator(leave_requests, 5)  # Show 5 requests per page
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    return render(request, 'hr/my_leave_request.html', {
        'user': user,
        'leave_requests': page_obj,
        'allocated_menus': allocated_menus,
    })


def leave_request_form(request):
    use = request.user  # Get the current user

    # Get user's permissions
    user_permissions = MenuPermissions.objects.filter(user=use).select_related('menu', 'sub_menu')

    # Create a dictionary to hold menus and their allocated submenus
    allocated_menus = {}
    for perm in user_permissions:
        if perm.menu not in allocated_menus:
            allocated_menus[perm.menu] = []
        allocated_menus[perm.menu].append(perm.sub_menu)

    user = request.user
    emp=EmployeeDetails.objects.get(user_id=user)
    if request.method == 'POST':
        form = LeaveRequestForm(request.POST)
        if form.is_valid():
            leave_request = form.save(commit=False)
            # Set the employee based on the logged-in user
            leave_request.employee_id =emp.id
            leave_request.save()
            return redirect('my_profile')  # Redirect after saving
    else:
        form = LeaveRequestForm()

    return render(request, 'hr/leave_request_form.html', {'form': form, 'allocated_menus': allocated_menus})


def search_approved_leave_requests(request):
    # Fetch all employees
    employees = EmployeeDetails.objects.all()

    # Prepare a list to hold employee leave requests
    leave_requests = []

    # Get search query from request
    search_query = request.GET.get('search_query', '').strip()

    for employee in employees:
        # Get all approved leave requests for each employee
        requests = RequestLeave.objects.filter(employee=employee, approval_status='Approved')

        print(requests)

        # If there is a search query, filter by employee name
        if search_query and search_query.lower() not in employee.user.name.lower():
            continue  # Skip this employee if the name does not match the search query

        leave_requests.append({
            'employee': employee,
            'requests': requests  # This will be a QuerySet of leave requests
        })

    # Apply pagination to leave_requests list
    paginator = Paginator(leave_requests, 10)  # Show 10 employees per page
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    use = request.user  # Get the current user

    # Get user's permissions
    user_permissions = MenuPermissions.objects.filter(user=use).select_related('menu', 'sub_menu')

    # Create a dictionary to hold menus and their allocated submenus
    allocated_menus = {}
    for perm in user_permissions:
        if perm.menu not in allocated_menus:
            allocated_menus[perm.menu] = []
        allocated_menus[perm.menu].append(perm.sub_menu)

    # Pass the paginated leave_requests and menus to the context
    context = {
        'allocated_menus': allocated_menus,
        'page_obj': page_obj,  # Include the page object in the context
        'search_query': search_query,  # Include search query for form repopulation
    }

    return render(request, 'hr/leave_list.html', context)

def employee_detail(request):
    use = request.user  # Get the current user

    # Get user's permissions
    user_permissions = MenuPermissions.objects.filter(user=use).select_related('menu', 'sub_menu')

    # Create a dictionary to hold menus and their allocated submenus
    allocated_menus = {}
    for perm in user_permissions:
        if perm.menu not in allocated_menus:
            allocated_menus[perm.menu] = []
        allocated_menus[perm.menu].append(perm.sub_menu)

    # Get initial queryset for finished goods based on the provided ID
    Employee = EmployeeDetails.objects.all()

    # Get filter type and query date from the request
    filter_type = request.GET.get('filter')
    query_date = request.GET.get('query')  # Expected format: YYYY-MM-DD
    is_filtered = False

    # Apply filtering if filter type is 'date' and query_date is provided
    if filter_type == 'date' and query_date:
        employee_list = Employee.filter(date=query_date)
        is_filtered = True  # Flag to indicate that a filter has been applied

    # Prepare context for rendering template
    context = {
        'data': Employee,
        'allocated_menus': allocated_menus,
        'filter_type': filter_type,
        'query_date': query_date,
        'is_filtered': is_filtered  # Pass flag to template
    }

    return render(request, 'hr/employee_detail.html', context)

def leave_history(request,id):
    use = request.user  # Get the current user

    # Get user's permissions
    user_permissions = MenuPermissions.objects.filter(user=use).select_related('menu', 'sub_menu')

    # Create a dictionary to hold menus and their allocated submenus
    allocated_menus = {}
    for perm in user_permissions:
        if perm.menu not in allocated_menus:
            allocated_menus[perm.menu] = []
        allocated_menus[perm.menu].append(perm.sub_menu)

    leave_request = RequestLeave.objects.filter(employee__user_id=id).order_by('-start_date')

    # Get filter type and query date from the request
    filter_type = request.GET.get('filter')
    query_date = request.GET.get('query')  # Expected format: YYYY-MM-DD
    is_filtered = False

    # Apply filtering if filter type is 'date' and query_date is provided
    if filter_type == 'date' and query_date:
        finished_good = leave_request.filter(date=query_date)
        is_filtered = True  # Flag to indicate that a filter has been applied

    # Prepare context for rendering template
    context = {
        'data': leave_request,
        'allocated_menus': allocated_menus,
        'filter_type': filter_type,
        'query_date': query_date,
        'is_filtered': is_filtered  # Pass flag to template
    }

    return render(request, 'hr/leave_history.html', context)


def create_payroll(request, id):
    # Retrieve the employee instance based on the given user ID
    employee = get_object_or_404(EmployeeDetails, user_id=id)
    basic_salary = employee.basic or Decimal('0.00')  # Ensure fallback to 0 if basic salary is null

    if request.method == 'POST':
        form = PayrollForm(request.POST)
        if form.is_valid():

            payroll_date = form.cleaned_data['pay_date']

            # Check if a payroll entry already exists for the same employee in the same month and year
            existing_payroll = Payroll.objects.filter(
                employee=employee,
                pay_date__year=payroll_date.year,
                pay_date__month=payroll_date.month
            ).exists()

            if existing_payroll:
                messages.warning(request,f"A payroll entry for {employee.user.name} already exists for {payroll_date.strftime('%B %Y')}.")
                return render(request, 'hr/payroll_form.html',{'form': form, 'employee': employee, 'basic_salary': basic_salary})

            payroll = form.save(commit=False)
            payroll.employee = employee  # Set the employee field for Payroll

            # Calculate gross earnings
            gross_earnings = (
                    basic_salary +
                    form.cleaned_data.get('hra', Decimal('0.00')) +
                    form.cleaned_data.get('da', Decimal('0.00')) +
                    form.cleaned_data.get('ot', Decimal('0.00')) +
                    form.cleaned_data.get('bonuses', Decimal('0.00'))+
                    form.cleaned_data.get('medical',Decimal('0.00'))+
                    form.cleaned_data.get('insurance',Decimal('0.00'))+
                    form.cleaned_data.get('ta', Decimal('0.00'))+
                    form.cleaned_data.get('food_allowance', Decimal('0.00'))
            )

            # Calculate total deductions
            total_deductions = (
                    form.cleaned_data.get('canteen_expense', Decimal('0.00')) +
                    form.cleaned_data.get('pf', Decimal('0.00')) +
                    form.cleaned_data.get('esi', Decimal('0.00')) +
                    form.cleaned_data.get('leave_plus', Decimal('0.00'))
            )

            # Calculate net pay and set it on the payroll instance
            payroll.net_pay = gross_earnings - total_deductions

            # Save the payroll instance with the calculated net pay
            payroll.save()

            # Redirect to the desired page after successful save
            return redirect('Employee_list')
    else:
        # Provide the basic salary to the form if needed
        form = PayrollForm()

    # Render the form along with basic salary and employee information
    return render(request, 'hr/payroll_form.html', {'form': form, 'employee': employee, 'basic_salary': basic_salary})

def payroll_summary(request,id):
    use = request.user  # Get the current user

    # Get user's permissions
    user_permissions = MenuPermissions.objects.filter(user=use).select_related('menu', 'sub_menu')

    # Create a dictionary to hold menus and their allocated submenus
    allocated_menus = {}
    for perm in user_permissions:
        if perm.menu not in allocated_menus:
            allocated_menus[perm.menu] = []
        allocated_menus[perm.menu].append(perm.sub_menu)

    payroll = Payroll.objects.filter(employee__user_id=id).order_by('-pay_date')

    # Get filter type and query date from the request
    filter_type = request.GET.get('filter')
    query_date = request.GET.get('query')  # Expected format: YYYY-MM-DD
    is_filtered = False

    # Apply filtering if filter type is 'date' and query_date is provided
    if filter_type == 'date' and query_date:
        finished_good = payroll.filter(date=query_date)
        is_filtered = True  # Flag to indicate that a filter has been applied

    # Prepare context for rendering template
    context = {
        'data': payroll,
        'allocated_menus': allocated_menus,
        'filter_type': filter_type,
        'query_date': query_date,
        'is_filtered': is_filtered  # Pass flag to template
    }

    return render(request, 'hr/payroll_summary.html', context)


def payroll_by_month(request):
    # Get current month and year
    current_month = timezone.now().month
    current_year = timezone.now().year

    # Get month and year from GET parameters, defaulting to current values
    selected_month = request.GET.get('month', current_month)
    selected_year = request.GET.get('year', current_year)

    # Convert selected_month and selected_year to integers
    try:
        selected_month = int(selected_month)
        selected_year = int(selected_year)
    except ValueError:
        selected_month = current_month
        selected_year = current_year

    # Fetch payroll records for the selected month and year
    payroll_records = Payroll.objects.filter(pay_date__month=selected_month,
                                             pay_date__year=selected_year)

    if not payroll_records.exists():
        messages.warning(request, 'No payroll records found for this month.')

    total_net_pay = sum(record.net_pay for record in payroll_records)

    # Define months as tuples (month_number, month_name)
    months = [
        (1, 'January'), (2, 'February'), (3, 'March'), (4, 'April'),
        (5, 'May'), (6, 'June'), (7, 'July'), (8, 'August'),
        (9, 'September'), (10, 'October'), (11, 'November'), (12, 'December')
    ]

    context = {
        'payrolls': payroll_records,
        'total_net_pay': total_net_pay,
        'months': months,
        'selected_month': selected_month,
        'selected_year': selected_year,
    }

    return render(request, 'hr/salary_list_by_month.html', context)